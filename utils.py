import os
import pandas as pd
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import importlib.util
import sys

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    """Check if file has allowed extension"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def load_measure_script(measure_number):
    """Dynamically load and return the measure processing function"""
    try:
        script_path = os.path.join('measures', f'{measure_number}.py')
        if not os.path.exists(script_path):
            raise FileNotFoundError(f"Measure script {measure_number}.py not found")
        
        spec = importlib.util.spec_from_file_location(f"measure_{measure_number}", script_path)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        
        # Look for the main processing function
        if hasattr(module, 'filter_patients'):
            return module.filter_patients
        elif hasattr(module, 'process_measure'):
            return module.process_measure
        else:
            raise AttributeError(f"No suitable function found in measure {measure_number}")
    
    except Exception as e:
        logging.error(f"Error loading measure {measure_number}: {str(e)}")
        raise

def process_excel_file(filepath, selected_measures, download_folder):
    """
    Process the uploaded Excel file with selected measures
    Returns: dict with success status and either download_path or error message
    """
    try:
        # Read the original Excel file
        logging.info(f"Reading Excel file: {filepath}")
        df = pd.read_excel(filepath)
        
        if df.empty:
            return {'success': False, 'error': 'The uploaded file is empty'}
        
        # Create a new workbook for results
        wb = Workbook()
        
        # Remove the default sheet
        wb.remove(wb.active)
        
        # Add original data sheet
        original_sheet = wb.create_sheet("Original Data")
        for row in dataframe_to_rows(df, index=False, header=True):
            original_sheet.append(row)
        
        # Process each selected measure
        summary_data = []
        
        for measure in selected_measures:
            try:
                logging.info(f"Processing measure {measure}")
                
                # Load the measure processing function
                filter_function = load_measure_script(measure)
                
                # Apply the filter function to the data
                filtered_df = filter_function(df.copy())
                
                if not filtered_df.empty:
                    # Create sheet for this measure
                    measure_sheet = wb.create_sheet(f"Measure {measure}")
                    for row in dataframe_to_rows(filtered_df, index=False, header=True):
                        measure_sheet.append(row)
                    
                    # Add to summary
                    summary_data.append({
                        'Measure': f"Measure {measure}",
                        'Eligible Patients': len(filtered_df),
                        'Total Patients': len(df)
                    })
                else:
                    # Create empty sheet with note
                    measure_sheet = wb.create_sheet(f"Measure {measure}")
                    measure_sheet.append(['No eligible patients found for this measure'])
                    
                    summary_data.append({
                        'Measure': f"Measure {measure}",
                        'Eligible Patients': 0,
                        'Total Patients': len(df)
                    })
                
            except Exception as e:
                logging.error(f"Error processing measure {measure}: {str(e)}")
                # Create error sheet
                error_sheet = wb.create_sheet(f"Measure {measure} - Error")
                error_sheet.append(['Error processing this measure:', str(e)])
                
                summary_data.append({
                    'Measure': f"Measure {measure}",
                    'Eligible Patients': 'Error',
                    'Total Patients': len(df)
                })
        
        # Add summary sheet
        if summary_data:
            summary_sheet = wb.create_sheet("Summary")
            summary_df = pd.DataFrame(summary_data)
            for row in dataframe_to_rows(summary_df, index=False, header=True):
                summary_sheet.append(row)
        
        # Save the workbook
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"processed_mips_report_{timestamp}.xlsx"
        download_path = os.path.join(download_folder, filename)
        
        wb.save(download_path)
        logging.info(f"Saved processed file to: {download_path}")
        
        return {
            'success': True,
            'download_path': download_path,
            'summary': summary_data
        }
        
    except Exception as e:
        logging.error(f"Error processing Excel file: {str(e)}")
        return {
            'success': False,
            'error': f"Processing failed: {str(e)}"
        }

def cleanup_old_files(folder_path, max_age_hours=24):
    """Clean up old files from upload/download folders"""
    try:
        current_time = datetime.now()
        for filename in os.listdir(folder_path):
            filepath = os.path.join(folder_path, filename)
            if os.path.isfile(filepath):
                file_time = datetime.fromtimestamp(os.path.getctime(filepath))
                age_hours = (current_time - file_time).total_seconds() / 3600
                
                if age_hours > max_age_hours:
                    os.remove(filepath)
                    logging.info(f"Cleaned up old file: {filepath}")
    except Exception as e:
        logging.error(f"Error during cleanup: {str(e)}")
